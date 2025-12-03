"""
Excel Report Generation for Finance Flow
Creates various Excel reports for budget analysis
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db.models import Sum, Q
from .models import Budget, Category, Transaction, Household
from datetime import datetime, date
import calendar


def get_header_style():
    """Returns style for header cells"""
    return {
        'font': Font(bold=True, color='FFFFFF', size=11),
        'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }


def get_subheader_style():
    """Returns style for subheader cells"""
    return {
        'font': Font(bold=True, size=10),
        'fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
        'alignment': Alignment(horizontal='left', vertical='center'),
    }


def apply_style(cell, style_dict):
    """Apply style dictionary to a cell"""
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def format_currency(value):
    """Format value as currency"""
    if value is None:
        return 0.00
    return float(value)


def export_yearly_budget(household, year):
    """Export yearly budget overview to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Budget {year}"
    
    # Title
    ws.merge_cells('A1:O1')
    title_cell = ws['A1']
    title_cell.value = f"Budget Overview - {year}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Month headers
    month_names = [calendar.month_abbr[i] for i in range(1, 13)]
    headers = ['Category', 'Type'] + month_names + ['Total', 'Average']
    
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, get_header_style())
        ws.column_dimensions[get_column_letter(col)].width = 12
    
    ws.column_dimensions['A'].width = 30  # Category name
    ws.column_dimensions['B'].width = 12  # Type
    
    # Get all categories for this household
    categories = Category.objects.filter(household=household).order_by('type', 'name')
    
    row = 4
    current_type = None
    
    for category in categories:
        # Add type header if new type
        if category.type != current_type:
            if current_type is not None:
                row += 1
            current_type = category.type
            type_cell = ws.cell(row=row, column=1)
            type_cell.value = category.get_type_display()
            apply_style(type_cell, get_subheader_style())
            ws.merge_cells(f'A{row}:O{row}')
            row += 1
        
        # Category name
        indent = "  " if category.parent else ""
        ws.cell(row=row, column=1).value = f"{indent}{category.name}"
        ws.cell(row=row, column=2).value = category.get_type_display()
        
        # Monthly amounts
        total = 0
        for month in range(1, 13):
            month_date = date(year, month, 1)
            budget = Budget.objects.filter(
                category=category,
                start_date__year=year,
                start_date__month=month
            ).first()
            
            amount = format_currency(budget.amount if budget else 0)
            cell = ws.cell(row=row, column=month + 2)
            cell.value = amount
            cell.number_format = '#,##0.00'
            total += amount
        
        # Total
        total_cell = ws.cell(row=row, column=15)
        total_cell.value = total
        total_cell.number_format = '#,##0.00'
        total_cell.font = Font(bold=True)
        
        # Average
        avg_cell = ws.cell(row=row, column=16)
        avg_cell.value = f"=O{row}/12"
        avg_cell.number_format = '#,##0.00'
        avg_cell.font = Font(bold=True)
        
        row += 1
    
    # Summary row
    row += 1
    summary_cell = ws.cell(row=row, column=1)
    summary_cell.value = "TOTALS"
    summary_cell.font = Font(bold=True, size=12)
    apply_style(summary_cell, get_subheader_style())
    
    # Sum formulas for each month
    for month in range(1, 13):
        col = month + 2
        sum_cell = ws.cell(row=row, column=col)
        sum_cell.value = f"=SUM({get_column_letter(col)}4:{get_column_letter(col)}{row-1})"
        sum_cell.number_format = '#,##0.00'
        sum_cell.font = Font(bold=True)
        apply_style(sum_cell, get_subheader_style())
    
    # Total column sum
    total_sum_cell = ws.cell(row=row, column=15)
    total_sum_cell.value = f"=SUM(O4:O{row-1})"
    total_sum_cell.number_format = '#,##0.00'
    total_sum_cell.font = Font(bold=True)
    apply_style(total_sum_cell, get_subheader_style())
    
    # Average column
    avg_sum_cell = ws.cell(row=row, column=16)
    avg_sum_cell.value = f"=O{row}/12"
    avg_sum_cell.number_format = '#,##0.00'
    avg_sum_cell.font = Font(bold=True)
    apply_style(avg_sum_cell, get_subheader_style())
    
    # Freeze panes
    ws.freeze_panes = 'C4'
    
    return wb


def export_monthly_detail(household, year, month):
    """Export detailed monthly budget to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"{calendar.month_abbr[month]} {year}"
    
    month_date = date(year, month, 1)
    month_name = calendar.month_name[month]
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Budget Detail - {month_name} {year}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Headers
    headers = ['Category', 'Type', 'Budgeted', 'Paid', 'Remaining', 'Status']
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, get_header_style())
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    ws.column_dimensions['A'].width = 30
    
    # Get categories and budgets
    categories = Category.objects.filter(household=household).order_by('type', 'name')
    
    row = 4
    current_type = None
    income_total = expense_total = savings_total = 0
    income_paid = expense_paid = savings_paid = 0
    
    for category in categories:
        if category.type != current_type:
            if current_type is not None:
                row += 1
            current_type = category.type
            type_cell = ws.cell(row=row, column=1)
            type_cell.value = category.get_type_display()
            apply_style(type_cell, get_subheader_style())
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
        
        budget = Budget.objects.filter(
            category=category,
            start_date__year=year,
            start_date__month=month
        ).first()
        
        budgeted = format_currency(budget.amount if budget else 0)
        paid = format_currency(budget.amount if (budget and budget.is_paid) else 0)
        remaining = budgeted - paid
        
        indent = "  " if category.parent else ""
        ws.cell(row=row, column=1).value = f"{indent}{category.name}"
        ws.cell(row=row, column=2).value = category.get_type_display()
        
        budget_cell = ws.cell(row=row, column=3)
        budget_cell.value = budgeted
        budget_cell.number_format = '#,##0.00'
        
        paid_cell = ws.cell(row=row, column=4)
        paid_cell.value = paid
        paid_cell.number_format = '#,##0.00'
        
        remaining_cell = ws.cell(row=row, column=5)
        remaining_cell.value = remaining
        remaining_cell.number_format = '#,##0.00'
        if remaining < 0:
            remaining_cell.font = Font(color='FF0000', bold=True)
        
        status_cell = ws.cell(row=row, column=6)
        if budget and budget.is_paid:
            status_cell.value = "Paid"
            status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif budgeted > 0:
            status_cell.value = "Pending"
            status_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        else:
            status_cell.value = "Not Budgeted"
        
        # Track totals by type
        if category.type == 'INCOME':
            income_total += budgeted
            income_paid += paid
        elif category.type == 'EXPENSE':
            expense_total += budgeted
            expense_paid += paid
        elif category.type == 'SAVINGS':
            savings_total += budgeted
            savings_paid += paid
        
        row += 1
    
    # Summary section
    row += 2
    summary_title = ws.cell(row=row, column=1)
    summary_title.value = "SUMMARY"
    summary_title.font = Font(bold=True, size=12)
    apply_style(summary_title, get_subheader_style())
    ws.merge_cells(f'A{row}:F{row}')
    row += 1
    
    summary_headers = ['Type', 'Budgeted', 'Paid', 'Remaining', 'Percentage']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, get_header_style())
    row += 1
    
    # Income summary
    ws.cell(row=row, column=1).value = "Income"
    ws.cell(row=row, column=2).value = income_total
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=3).value = income_paid
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=4).value = income_total - income_paid
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    ws.cell(row=row, column=5).value = f"=C{row}/B{row}*100" if income_total > 0 else 0
    ws.cell(row=row, column=5).number_format = '0.0"%"'
    row += 1
    
    # Expense summary
    ws.cell(row=row, column=1).value = "Expenses"
    ws.cell(row=row, column=2).value = expense_total
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=3).value = expense_paid
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=4).value = expense_total - expense_paid
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    ws.cell(row=row, column=5).value = f"=C{row}/B{row}*100" if expense_total > 0 else 0
    ws.cell(row=row, column=5).number_format = '0.0"%"'
    row += 1
    
    # Savings summary
    ws.cell(row=row, column=1).value = "Savings"
    ws.cell(row=row, column=2).value = savings_total
    ws.cell(row=row, column=2).number_format = '#,##0.00'
    ws.cell(row=row, column=3).value = savings_paid
    ws.cell(row=row, column=3).number_format = '#,##0.00'
    ws.cell(row=row, column=4).value = savings_total - savings_paid
    ws.cell(row=row, column=4).number_format = '#,##0.00'
    ws.cell(row=row, column=5).value = f"=C{row}/B{row}*100" if savings_total > 0 else 0
    ws.cell(row=row, column=5).number_format = '0.0"%"'
    row += 1
    
    # Net balance
    row += 1
    balance_cell = ws.cell(row=row, column=1)
    balance_cell.value = "NET BALANCE"
    balance_cell.font = Font(bold=True, size=11)
    balance_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    balance_amount = (income_total - income_paid) - (expense_total - expense_paid) - (savings_total - savings_paid)
    balance_value = ws.cell(row=row, column=4)
    balance_value.value = balance_amount
    balance_value.number_format = '#,##0.00'
    balance_value.font = Font(bold=True, size=11)
    if balance_amount < 0:
        balance_value.font = Font(bold=True, size=11, color='FF0000')
    balance_value.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    return wb


def export_category_summary(household, year):
    """Export category summary with yearly totals"""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Category Summary {year}"
    
    # Title
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = f"Category Summary - {year}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Headers
    headers = ['Category', 'Type', 'Total Budgeted', 'Average Monthly', 'Months Active']
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, get_header_style())
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # Get categories
    categories = Category.objects.filter(household=household).order_by('type', 'name')
    
    row = 4
    current_type = None
    
    for category in categories:
        if category.type != current_type:
            if current_type is not None:
                row += 1
            current_type = category.type
            type_cell = ws.cell(row=row, column=1)
            type_cell.value = category.get_type_display()
            apply_style(type_cell, get_subheader_style())
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
        
        # Calculate totals
        budgets = Budget.objects.filter(
            category=category,
            start_date__year=year
        )
        
        total = sum(format_currency(b.amount) for b in budgets)
        months_active = budgets.count()
        average = total / 12 if months_active > 0 else 0
        
        indent = "  " if category.parent else ""
        ws.cell(row=row, column=1).value = f"{indent}{category.name}"
        ws.cell(row=row, column=2).value = category.get_type_display()
        
        total_cell = ws.cell(row=row, column=3)
        total_cell.value = total
        total_cell.number_format = '#,##0.00'
        
        avg_cell = ws.cell(row=row, column=4)
        avg_cell.value = average
        avg_cell.number_format = '#,##0.00'
        
        active_cell = ws.cell(row=row, column=5)
        active_cell.value = months_active
        active_cell.alignment = Alignment(horizontal='center')
        
        row += 1
    
    return wb


def export_transactions(household, start_date=None, end_date=None):
    """Export transaction history to Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Title
    date_range = ""
    if start_date and end_date:
        date_range = f" ({start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')})"
    elif start_date:
        date_range = f" (from {start_date.strftime('%Y-%m-%d')})"
    elif end_date:
        date_range = f" (until {end_date.strftime('%Y-%m-%d')})"
    
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"Transaction History{date_range}"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Headers
    headers = ['Date', 'Category', 'Type', 'Description', 'Amount', 'Balance']
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, get_header_style())
    
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Get transactions
    transactions = Transaction.objects.filter(household=household)
    if start_date:
        transactions = transactions.filter(date__gte=start_date)
    if end_date:
        transactions = transactions.filter(date__lte=end_date)
    transactions = transactions.order_by('date', 'id')
    
    row = 4
    running_balance = 0
    
    for transaction in transactions:
        if transaction.type == 'INCOME':
            running_balance += format_currency(transaction.amount)
        else:
            running_balance -= format_currency(transaction.amount)
        
        ws.cell(row=row, column=1).value = transaction.date
        ws.cell(row=row, column=1).number_format = 'YYYY-MM-DD'
        
        category_name = transaction.category.name if transaction.category else "Uncategorized"
        ws.cell(row=row, column=2).value = category_name
        ws.cell(row=row, column=3).value = transaction.get_type_display()
        ws.cell(row=row, column=4).value = transaction.description
        
        amount_cell = ws.cell(row=row, column=5)
        amount_cell.value = format_currency(transaction.amount)
        amount_cell.number_format = '#,##0.00'
        if transaction.type == 'INCOME':
            amount_cell.font = Font(color='006100')
        else:
            amount_cell.font = Font(color='C00000')
        
        balance_cell = ws.cell(row=row, column=6)
        balance_cell.value = running_balance
        balance_cell.number_format = '#,##0.00'
        if running_balance < 0:
            balance_cell.font = Font(color='FF0000', bold=True)
        
        row += 1
    
    # Summary
    if transactions.exists():
        row += 1
        summary_cell = ws.cell(row=row, column=1)
        summary_cell.value = "TOTALS"
        summary_cell.font = Font(bold=True, size=12)
        apply_style(summary_cell, get_subheader_style())
        
        income_total = transactions.filter(type='INCOME').aggregate(Sum('amount'))['amount__sum'] or 0
        expense_total = transactions.filter(type='EXPENSE').aggregate(Sum('amount'))['amount__sum'] or 0
        
        ws.cell(row=row, column=2).value = "Income"
        ws.cell(row=row, column=5).value = format_currency(income_total)
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=5).font = Font(bold=True, color='006100')
        
        row += 1
        ws.cell(row=row, column=2).value = "Expenses"
        ws.cell(row=row, column=5).value = format_currency(expense_total)
        ws.cell(row=row, column=5).number_format = '#,##0.00'
        ws.cell(row=row, column=5).font = Font(bold=True, color='C00000')
        
        row += 1
        ws.cell(row=row, column=2).value = "Net"
        net_cell = ws.cell(row=row, column=5)
        net_cell.value = format_currency(income_total - expense_total)
        net_cell.number_format = '#,##0.00'
        net_cell.font = Font(bold=True, size=11)
        if income_total - expense_total < 0:
            net_cell.font = Font(bold=True, size=11, color='FF0000')
    
    return wb


def export_category_setup(household):
    """Export category setup information with pivot structure by main category"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Category Setup"
    
    # Title
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "Category Setup Information"
    title_cell.font = Font(bold=True, size=16)
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 25
    
    # Headers
    headers = ['Category', 'Type', 'Main Category', 'Persistent', 'Payment Type', 'Essential']
    row = 3
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        apply_style(cell, get_header_style())
    
    # Column widths
    ws.column_dimensions['A'].width = 35  # Category name
    ws.column_dimensions['B'].width = 15  # Type
    ws.column_dimensions['C'].width = 25  # Main Category
    ws.column_dimensions['D'].width = 12  # Persistent
    ws.column_dimensions['E'].width = 25  # Payment Type
    ws.column_dimensions['F'].width = 12  # Essential
    
    # Get all categories grouped by type
    categories_by_type = {}
    all_categories = Category.objects.filter(household=household).select_related('parent').prefetch_related('children')
    
    for category in all_categories:
        if category.type not in categories_by_type:
            categories_by_type[category.type] = []
        categories_by_type[category.type].append(category)
    
    row = 4
    current_type = None
    
    # Process each type
    for cat_type in ['INCOME', 'EXPENSE', 'SAVINGS']:
        if cat_type not in categories_by_type:
            continue
            
        # Add type header
        if current_type is not None:
            row += 1
        current_type = cat_type
        type_cell = ws.cell(row=row, column=1)
        # Get type display name from first category of this type
        if categories_by_type[cat_type]:
            type_display = categories_by_type[cat_type][0].get_type_display()
        else:
            type_display = {'INCOME': 'Income', 'EXPENSE': 'Expense', 'SAVINGS': 'Savings & Investments'}.get(cat_type, cat_type)
        type_cell.value = type_display
        apply_style(type_cell, get_subheader_style())
        ws.merge_cells(f'A{row}:F{row}')
        row += 1
        
        # Separate main categories and sub-categories
        main_cats = [c for c in categories_by_type[cat_type] if c.parent is None]
        sub_cats = [c for c in categories_by_type[cat_type] if c.parent is not None]
        
        # Sort main categories by name
        main_cats.sort(key=lambda x: x.name)
        
        # Process each main category with its sub-categories
        for main_cat in main_cats:
            # Create main category row (bold, highlighted)
            main_cell = ws.cell(row=row, column=1)
            main_cell.value = main_cat.name
            main_cell.font = Font(bold=True, size=11)
            main_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
            type_cell = ws.cell(row=row, column=2)
            type_cell.value = main_cat.get_type_display()
            type_cell.font = Font(bold=True, size=11)
            type_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
            main_cat_cell = ws.cell(row=row, column=3)
            main_cat_cell.value = main_cat.name
            main_cat_cell.font = Font(bold=True, size=11)
            main_cat_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
            persistent_cell = ws.cell(row=row, column=4)
            persistent_cell.value = "Yes" if main_cat.is_persistent else "No"
            persistent_cell.font = Font(bold=True, size=11)
            persistent_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            persistent_cell.alignment = Alignment(horizontal='center')
            
            payment_cell = ws.cell(row=row, column=5)
            payment_cell.value = main_cat.get_payment_type_display()
            payment_cell.font = Font(bold=True, size=11)
            payment_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
            essential_cell = ws.cell(row=row, column=6)
            essential_cell.value = "Yes" if main_cat.is_essential else "No"
            essential_cell.font = Font(bold=True, size=11)
            essential_cell.fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            essential_cell.alignment = Alignment(horizontal='center')
            
            row += 1
            
            # Add sub-categories for this main category
            sub_cats_for_main = [c for c in sub_cats if c.parent_id == main_cat.id]
            sub_cats_for_main.sort(key=lambda x: x.name)
            
            for sub_cat in sub_cats_for_main:
                # Category name (indented)
                sub_cell = ws.cell(row=row, column=1)
                sub_cell.value = f"  └─ {sub_cat.name}"
                sub_cell.font = Font(size=10)
                
                type_cell = ws.cell(row=row, column=2)
                type_cell.value = sub_cat.get_type_display()
                
                main_cat_cell = ws.cell(row=row, column=3)
                main_cat_cell.value = main_cat.name
                
                persistent_cell = ws.cell(row=row, column=4)
                persistent_cell.value = "Yes" if sub_cat.is_persistent else "No"
                persistent_cell.alignment = Alignment(horizontal='center')
                
                payment_cell = ws.cell(row=row, column=5)
                payment_cell.value = sub_cat.get_payment_type_display()
                
                essential_cell = ws.cell(row=row, column=6)
                essential_cell.value = "Yes" if sub_cat.is_essential else "No"
                essential_cell.alignment = Alignment(horizontal='center')
                
                row += 1
    
    # Add summary section
    row += 2
    summary_row = row
    
    ws.merge_cells(f'A{summary_row}:F{summary_row}')
    summary_cell = ws.cell(row=summary_row, column=1)
    summary_cell.value = "Summary"
    summary_cell.font = Font(bold=True, size=14)
    summary_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    summary_cell.font = Font(bold=True, color='FFFFFF', size=14)
    summary_cell.alignment = Alignment(horizontal='center')
    row += 1
    
    # Count totals
    total_categories = Category.objects.filter(household=household).count()
    main_categories = Category.objects.filter(household=household, parent__isnull=True).count()
    sub_categories = Category.objects.filter(household=household, parent__isnull=False).count()
    persistent_count = Category.objects.filter(household=household, is_persistent=True).count()
    manual_tracking = Category.objects.filter(household=household, payment_type='MANUAL').count()
    essential_count = Category.objects.filter(household=household, is_essential=True).count()
    
    summary_data = [
        ['Total Categories', total_categories],
        ['Main Categories', main_categories],
        ['Sub-Categories', sub_categories],
        ['Persistent Categories', persistent_count],
        ['Manual Tracking Categories', manual_tracking],
        ['Essential Categories', essential_count],
    ]
    
    for label, value in summary_data:
        label_cell = ws.cell(row=row, column=1)
        label_cell.value = label
        label_cell.font = Font(bold=True)
        
        value_cell = ws.cell(row=row, column=2)
        value_cell.value = value
        value_cell.alignment = Alignment(horizontal='center')
        
        row += 1
    
    return wb

